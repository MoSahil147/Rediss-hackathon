import os
import json
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
import boto3
from dotenv import load_dotenv
from datetime import datetime

now = datetime.now()

year = now.year
month = now.month

# Load credentials from .env
load_dotenv(dotenv_path=".env")

AWS_ACCESS_KEY = os.getenv("AWS_ACCESS_KEY_ID_EXCEL_TEMPLATE")
AWS_SECRET_KEY = os.getenv("AWS_SECRET_ACCESS_KEY_EXCEL_TEMPLATE")
AWS_REGION = os.getenv("AWS_DEFAULT_REGION_EXCEL_TEMPLATE", "ap-south-1")
S3_BUCKET = os.getenv("BUCKET_NAME_EXCEL_TEMPLATE")

# --- Utility Functions ---

def normalize(text: str) -> str:
    """Removes spaces, underscores, and makes text lowercase for matching."""
    return str(text).strip().lower().replace(" ", "").replace("_", "")


def extend_dropdowns(ws, max_row_written):
    """
    Extend all data validation ranges in a worksheet down to `max_row_written`.
    """
    if not ws.data_validations.dataValidation:
        return  # No dropdowns to extend

    for dv in ws.data_validations.dataValidation:
        new_sqrefs = []
        for sqref in dv.sqref.ranges:
            min_col, min_row, max_col, _ = range_boundaries(str(sqref))
            new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row_written}"
            new_sqrefs.append(new_range)
        dv.sqref = " ".join(new_sqrefs)


# def write_to_sheet(sheet_name, df, wb):
#     print(f"\nProcessing sheet: '{sheet_name}'...")
#     if sheet_name not in wb.sheetnames:
#         print(f"  - ⚠️ WARNING: Sheet '{sheet_name}' not found in the template. Skipping.")
#         return
#     ws = wb[sheet_name]

#     # Read headers from Row 1 (the orange row)
#     headers = [cell.value for cell in ws[1] if cell.value]
#     print(f"  - Reading template headers from Row 1: {headers}")

#     # Build header mapping
#     mapping = {}
#     norm_json_cols = {normalize(col): col for col in df.columns}
#     for h in headers:
#         norm_h = normalize(h)
#         if norm_h in norm_json_cols:
#             mapping[h] = norm_json_cols[norm_h]

#     if not mapping:
#         print(f"  - ⚠️ CRITICAL: No matching headers found between JSON and template sheet '{sheet_name}'. No data will be written.")
#         return
#     print(f"  - Successfully mapped {len(mapping)} columns.")

#     # Data starts on Row 3, so we write from there.
#     max_row_written = 2
#     for r_idx, record in enumerate(df.to_dict(orient="records"), start=3):
#         for c_idx, h in enumerate(headers, start=1):
#             json_key = mapping.get(h)
#             if json_key:
#                 value = record.get(json_key)
#                 if value is not None and pd.notna(value):
#                     ws.cell(row=r_idx, column=c_idx, value=value)
#         max_row_written = r_idx
#     print(f"  - Wrote {max(0, max_row_written - 2)} rows of data.")

#     if max_row_written > 2:
#         extend_dropdowns(ws, max_row_written)
#         print("  - Extended dropdowns to cover new data.")


def write_to_sheet(sheet_name, df, wb):
    print(f"\nProcessing sheet: '{sheet_name}'...")
    if sheet_name not in wb.sheetnames:
        print(f"  - ⚠️ WARNING: Sheet '{sheet_name}' not found in the template. Skipping.")
        return
    ws = wb[sheet_name]

    # Read headers from Row 1 (the orange row)
    headers = [cell.value for cell in ws[1] if cell.value]
    print(f"  - Reading template headers from Row 1: {headers}")

    # Build header mapping
    mapping = {}
    norm_json_cols = {normalize(col): col for col in df.columns}
    for h in headers:
        norm_h = normalize(h)
        if norm_h in norm_json_cols:
            mapping[h] = norm_json_cols[norm_h]

    if not mapping:
        print(f"  - ⚠️ CRITICAL: No matching headers found between JSON and template sheet '{sheet_name}'. No data will be written.")
        return
    print(f"  - Successfully mapped {len(mapping)} columns.")

    # Data starts on Row 3, so we write from there.
    max_row_written = 2
    for r_idx, record in enumerate(df.to_dict(orient="records"), start=3):
        for c_idx, h in enumerate(headers, start=1):
            json_key = mapping.get(h)
            if json_key:
                value = record.get(json_key)
                if value is not None and pd.notna(value):
                    
                    # --- START: NEW DATE CONVERSION LOGIC ---
                    # Check if the current column header is 'BL Date' (or a normalized version)
                    # and the value is a string that can be converted.
                    if normalize(h) == normalize("BL Date") and isinstance(value, str):
                        try:
                            # Use pandas to_datetime which is flexible with formats
                            datetime_obj = pd.to_datetime(value)
                            # Write the datetime object to the cell
                            ws.cell(row=r_idx, column=c_idx, value=datetime_obj)
                        except (ValueError, TypeError):
                            # If conversion fails for any reason, write the original string
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    else:
                        # For all other columns, write the value as it is
                        ws.cell(row=r_idx, column=c_idx, value=value)
                    # --- END: NEW DATE CONVERSION LOGIC ---
                        
        max_row_written = r_idx
    print(f"  - Wrote {max(0, max_row_written - 2)} rows of data.")

    if max_row_written > 2:
        extend_dropdowns(ws, max_row_written)
        print("  - Extended dropdowns to cover new data.")
        

def excel_template_downlaoder(bucket_name=S3_BUCKET):
    """Upload a file to S3 using credentials from .env"""
    S3_KEY = "manifest/TEMPLATES/MPCI Bulk Excel.xlsx"
    s3_key = S3_KEY
    output_file = "Excel_Template.xlsx"
    try:
        s3 = boto3.client(
            "s3",
            aws_access_key_id=AWS_ACCESS_KEY,
            aws_secret_access_key=AWS_SECRET_KEY,
            region_name=AWS_REGION,
        )
        s3.download_file(bucket_name, S3_KEY, output_file)
        print(f"✅ File downloaded successfully as {output_file}")
        return f"{output_file}"
    except Exception as e:
        print(f"❌ Failed to Download Excel Template file: {e}")
        return None

def upload_to_s3(filename, file_path, bucket_name=S3_BUCKET):
    """Upload a file to S3 using credentials from .env"""
    # S3_KEY = f"manifest/MPCI_BULK/{year}/{month}/{filename}.xlsx"
    # Add .xlsx extension only if not already present
    if not filename.endswith('.xlsx'):
        filename_with_ext = f"{filename}.xlsx"
    else:
        filename_with_ext = filename
    S3_KEY = f"manifest/MPCI_BULK/{year}/{month}/{filename_with_ext}"
    s3_key = S3_KEY
    try:
        s3 = boto3.client(
            "s3",
            aws_access_key_id=AWS_ACCESS_KEY,
            aws_secret_access_key=AWS_SECRET_KEY,
            region_name=AWS_REGION,
        )
        s3.upload_file(file_path, bucket_name, s3_key)
        print(f"✅ File uploaded to s3://{bucket_name}/{s3_key}")
        return f"{bucket_name}/{s3_key}"
    except Exception as e:
        print(f"❌ Failed to upload file: {e}")
        return None
    
# --- Main Modular Pipeline ---
def process_json_to_excel(json_file, source, data_dict, template_file, output_file):
    # Step 1: Copy template
    shutil.copy(template_file, output_file)

    # Step 2: Load JSON
    try:
        with open(json_file, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"FATAL ERROR: The JSON file was not found at '{json_file}'")
        return
    except json.JSONDecodeError:
        print(f"FATAL ERROR: The file '{json_file}' is not a valid JSON. Check for syntax errors.")
        return

    # Create DataFrames from the JSON data
    dfs = {
        "BL Details": pd.DataFrame(data.get("blDetails", [])),
        "Container Details": pd.DataFrame(data.get("containerDetails", [])),
        "Item Details": pd.DataFrame(data.get("itemDetails", [])),
    }

     # --- START: NEW LOGIC ---
    # Add the Excel row number to each DataFrame before writing.
    # Data writing starts on row 3 in the template.
    if not dfs["BL Details"].empty:
        dfs["BL Details"]["blRow"] = dfs["BL Details"].index + 2

    if not dfs["Container Details"].empty:
        dfs["Container Details"]["cntRow"] = dfs["Container Details"].index + 2

    if not dfs["Item Details"].empty:
        dfs["Item Details"]["itmRow"] = dfs["Item Details"].index + 2
    # --- END: NEW LOGIC ---


    # --- Define the explicit column mappings needed ---
    item_details_mapping = {
        "cargoGrossWeight": "Cargo Gross Weight (KGM)",
        "cargoNetWeight": "Cargo Net Weight (KGM)",
        "marksAndNumbers": "Marks & Numbers"
    }

    all_mappings = {
        "BL Details": {}, # No mappings needed for this sheet
        "Container Details": {}, # No mappings needed for this sheet
        "Item Details": item_details_mapping,
    }

    #Mapping Fields from Frontend

    if source == "WEB_APP":
        # Helper function to check if a value is truthy and not empty/null
        def is_valid_value(val):
            if val is None:
                return False
            if isinstance(val, str):
                stripped = val.strip().lower()
                return stripped not in ["", "null", "none"]
            return bool(val)
        
        parent_bl_number = data_dict.get("parentBlNumber")
        if is_valid_value(parent_bl_number):
            dfs["BL Details"]["parentBlNumber"] = parent_bl_number
        vessel_name = data_dict.get("vesselName")
        if is_valid_value(vessel_name):
            dfs["BL Details"]["vesselName"] = vessel_name
        # voyage_id = data_dict.get("voyageId")
        # if is_valid_value(voyage_id):
        #     dfs["BL Details"]["voyageNumber"] = voyage_id
        # Map partyName from issuingPartyName; if empty, fallback to ffPartyName
        party_name = data_dict.get("issuingPartyName") or data_dict.get("ffPartyName")
        if is_valid_value(party_name):
            dfs["BL Details"]["partyName"] = party_name
        # Map partyMpciId from issuingPartyMpciId; if empty, fallback to ffPartyMpciId
        party_mpci_id = data_dict.get("issuingPartyMpciId") or data_dict.get("ffPartyMpciId")
        if is_valid_value(party_mpci_id):
            dfs["BL Details"]["partyMpciId"] = party_mpci_id
        mbl_issuing_party_name = data_dict.get("mblIssuingPartyName")
        if is_valid_value(mbl_issuing_party_name):
            dfs["BL Details"]["parentBlIssuingPartyName"] = mbl_issuing_party_name
        mbl_issuing_party_mpci_id = data_dict.get("mblIssuingPartyMpciId")
        if is_valid_value(mbl_issuing_party_mpci_id):
            dfs["BL Details"]["parentBlIssuingPartyMpciId"] = mbl_issuing_party_mpci_id
        # Do not override freightForwardingAgentName from frontend; preserve source JSON value


    # Step 3: Load the workbook
    wb = load_workbook(output_file)

    # Step 4: Rename columns where needed, then write each section to the workbook
    for sheet_name, df in dfs.items():
        if not df.empty:
            # Get the mapping for the current sheet
            mapping = all_mappings.get(sheet_name, {})
            
            # If a mapping exists, rename the columns before writing
            if mapping:
                df.rename(columns=mapping, inplace=True)
            
            # CORRECTED: Call the write function with the (now renamed) DataFrame
            write_to_sheet(sheet_name, df, wb)
        else:
            print(f"\nSkipping sheet '{sheet_name}' as there is no data in the JSON.")

    # Step 5: CORRECTED: Save the final workbook with all the written data
    wb.save(output_file)
    print(f"\n✅ Data export complete. Output file → {output_file}")
    return output_file, dfs

# def process_json_to_excel(json_file, template_file, output_file):
#     # Step 1: Copy template
#     shutil.copy(template_file, output_file)

#     # Step 2: Load JSON
#     try:
#         with open(json_file, "r", encoding="utf-8-sig") as f:
#             data = json.load(f)
#     except FileNotFoundError:
#         print(f"FATAL ERROR: The JSON file was not found at '{json_file}'")
#         return
#     except json.JSONDecodeError:
#         print(f"FATAL ERROR: The file '{json_file}' is not a valid JSON. Check for syntax errors.")
#         return

#     dfs = {
#         "BL Details": pd.DataFrame(data.get("blDetails", [])),
#         "Container Details": pd.DataFrame(data.get("containerDetails", [])),
#         "Item Details": pd.DataFrame(data.get("itemDetails", [])),
#     }

#     # Step 3: Load workbook
#     wb = load_workbook(output_file)

#     # Step 4: Write each section
#     for sheet_name, df in dfs.items():
#         if not df.empty:
#             write_to_sheet(sheet_name, df, wb)
#         else:
#             print(f"\nSkipping sheet '{sheet_name}' as there is no data in the JSON.")

#     # Step 5: Save output
#     wb.save(output_file)
#     print(f"\n✅ Data export complete. Output file → {output_file}")
#     return output_file


# --- Entry Point (for standalone run) ---

if __name__ == "__main__":
    json_file = "intermediate_json.json"
    template_file = "Excel_Template.xlsx"
    output_file = "Output_filled.xlsx"

    process_json_to_excel(json_file, template_file, output_file)
